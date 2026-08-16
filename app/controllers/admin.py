# Blueprint del rol administrador (RF-1.4: el admin gestiona cuentas).
# Incluye el módulo de gestión de conductores (RF-2: Gestión de flota).
# Es para exportar los datos de los conductores y vehículos a un archivo CSV
# Es para importar y exportar archivos en memoria, como el CSV que se genera para descargar los datos de los conductores y vehículos
# Módulo para normalizar cadenas de texto, útil para búsquedas insensibles a acentos y mayúsculas/minúsculas
# Herramienta para trabajar con fechas y horas
# Nos sirve para crear decoradores de funciones, como admin_required

import csv 
import io 
import math
import unicodedata 
from datetime import date, datetime, timedelta 
from functools import wraps 

from flask import Blueprint, Response, abort, flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import case
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload

from app import db
from app.models.alerta import Alerta
from app.models.bitacora import Bitacora
from app.models.conductor import Conductor
from app.models.emergencia import Emergencia
from app.models.reporte_averia import ReporteAveria
from app.models.usuario import Usuario
from app.models.vehiculo import Vehiculo
from app.models.viaje import Viaje
from app.controllers.perfil import procesar_perfil
from app.services.correo import enviar_correo
from app.utils import validar_datos_viaje

VIAJE_ESTADOS_CERRABLES_POR_ADMIN = ("activo", "alerta", "emergencia")

# Estados de viaje que cuentan como "sin cerrar" para efectos de no poder
# asignarle a un conductor o vehículo una ruta nueva (ver
# _conductores_disponibles_para_programar y el conflicto al guardar en
# viajes_programar).
VIAJE_ESTADOS_SIN_CERRAR = ("programado", "activo", "alerta", "emergencia")

# Historial de eventos: cada tipo normalizado con su etiqueta, color de badge
# (reutiliza el mismo sistema badge-status del resto del proyecto) y a qué
# lista existente apunta su link "Ver detalle". El proyecto no tiene vistas
# de detalle por registro individual, así que el link va a la lista general
# correspondiente: para Alertas eso significa que si la alerta ya está
# atendida no aparecerá ahí (esa lista solo muestra atendida=False), y para
# Viajes Activos significa que un viaje ya cerrado tampoco aparecerá (esa
# lista solo muestra activo/alerta/emergencia). Es una limitación conocida,
# no un bug: no hay a dónde más apuntar sin construir páginas de detalle
# nuevas, que quedan fuera del alcance pedido aquí.
TIPO_INFO_HISTORIAL = {
    "retraso": ("Alerta de retraso", "warning", "admin.alertas_lista"),
    "panico": ("Aviso de emergencia", "danger", "admin.alertas_lista"),
    "licencia_vencida": ("Licencia vencida", "danger", "admin.alertas_lista"),
    "asistencia_mecanica": ("Problema mecánico", "warning", "admin.alertas_lista"),
    "incidencia_trafico": ("Tráfico o retraso", "warning", "admin.alertas_lista"),
    "ruta_no_iniciada": ("Ruta no iniciada", "danger", "admin.alertas_lista"),
    "averia": ("Avería en ruta", "warning", "admin.reportes_averia_lista"),
    "cerrado_forzado": ("Viaje cerrado (forzado)", "danger", "admin.viajes_lista"),
    "completado": ("Viaje completado", "success", "admin.viajes_lista"),
}
TAM_PAGINA_HISTORIAL = 20

admin = Blueprint("admin", __name__)


def _normalizar_busqueda(texto):
    """Quita acentos y pasa a minúsculas para comparar en los buscadores de
    Conductores/Vehículos (?buscar=): sin esto, escribir "cesar" no
    encontraría a "César" y el buscador se sentiría roto para nombres en
    español, que casi siempre llevan acento.
    """
    sin_acentos = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return sin_acentos.lower()


def admin_required(view_func):
    """Exige sesión iniciada y rol admin (RF-1.3), igual que hacía dashboard()."""

    @wraps(view_func)
    @login_required
    def wrapper(*args, **kwargs):
        if not current_user.es_admin():
            flash("No tienes permiso para acceder a esa sección", "error")
            return redirect(url_for("auth.login"))
        return view_func(*args, **kwargs)

    return wrapper


def _datos_conductor_desde_formulario():
    """Lee y valida los campos del formulario de conductor.

    Retorna un dict listo para crear/actualizar un Conductor, o None (y ya
    hizo flash del error) si la fecha de vencimiento no es válida.
    """
    fecha_texto = request.form.get("fecha_vencimiento_lic", "")
    try:
        fecha_vencimiento = datetime.strptime(fecha_texto, "%Y-%m-%d").date()
    except ValueError:
        flash("La fecha de vencimiento de la licencia no es válida.", "error")
        return None

    return {
        "nombre": request.form.get("nombre", "").strip(),
        "telefono": request.form.get("telefono", "").strip(),
        "num_licencia": request.form.get("num_licencia", "").strip(),
        "fecha_vencimiento_lic": fecha_vencimiento,
        "contacto_emergencia": request.form.get("contacto_emergencia", "").strip(),
        "tel_emergencia": request.form.get("tel_emergencia", "").strip(),
    }


def _conductor_para_formulario(conductor):
    """Convierte un Conductor de BD a dict apto para prellenar el formulario HTML.

    El correo no vive en Conductor sino en la cuenta Usuario ligada a él
    (ver _guardar_email_cuenta), así que se busca aparte para prellenar el
    campo del formulario.
    """
    usuario = Usuario.query.filter_by(id_conductor=conductor.id_conductor).first()
    return {
        "nombre": conductor.nombre,
        "telefono": conductor.telefono,
        "num_licencia": conductor.num_licencia,
        "fecha_vencimiento_lic": conductor.fecha_vencimiento_lic.isoformat(),
        "contacto_emergencia": conductor.contacto_emergencia,
        "tel_emergencia": conductor.tel_emergencia,
        "email": usuario.email if usuario and usuario.email else "",
    }


def _guardar_email_cuenta(conductor, email):
    """Guarda `email` en la cuenta Usuario ligada a este conductor (RF-1: usado
    por "olvidé mi contraseña"). Solo se usa desde conductores_editar: desde
    conductores_nuevo el Conductor y su Usuario se crean juntos en la misma
    transacción, así que ese caso ya no puede darse ahí.

    Sigue haciendo falta en edición porque hay conductores creados antes de
    este cambio (o por seed.py) cuya cuenta Usuario, si existe, no quedó
    creada desde este formulario: si esa cuenta no existe todavía, el correo
    capturado no tiene dónde guardarse y se avisa al admin en vez de
    perderlo en silencio.
    """
    if not email:
        return

    usuario = Usuario.query.filter_by(id_conductor=conductor.id_conductor).first()
    if usuario is None:
        flash(
            "El correo no se guardó: este conductor todavía no tiene una cuenta de "
            "usuario vinculada. Vuelve a editarlo una vez que se le cree la cuenta.",
            "warning",
        )
        return

    usuario.email = email
    db.session.commit()


def _guardar_contrasena_cuenta(conductor, nueva_contrasena):
    """Actualiza la contraseña de la cuenta Usuario ligada a este conductor
    (mismo caso de uso que mecanicos_editar: única vía para ayudar a un
    conductor sin correo registrado que olvidó su contraseña).

    Mismo criterio que _guardar_email_cuenta: si el conductor todavía no
    tiene cuenta vinculada, no hay dónde guardarla y se avisa al admin en
    vez de perderla en silencio. Retorna True si sí se guardó, para que el
    llamador pueda reflejarlo en la Bitácora.
    """
    usuario = Usuario.query.filter_by(id_conductor=conductor.id_conductor).first()
    if usuario is None:
        flash(
            "La contraseña no se guardó: este conductor todavía no tiene una cuenta "
            "de usuario vinculada. Vuelve a editarlo una vez que se le cree la cuenta.",
            "warning",
        )
        return False

    usuario.set_password(nueva_contrasena)
    db.session.commit()
    return True


def _datos_cuenta_nueva_desde_formulario():
    """Lee y valida los campos de la cuenta de acceso en el alta de conductor
    (RF-1.1): nombre de usuario, contraseña temporal y correo. Los tres son
    obligatorios solo en este formulario de creación -el correo sigue siendo
    opcional a nivel de columna en BD, ver _guardar_email_cuenta para el caso
    de edición-, porque sin ellos la cuenta no podría usarse para iniciar
    sesión ni para "olvidé mi contraseña".

    Retorna un dict con nombre_usuario/contrasena/email, o None (y ya hizo
    flash del error) si falta algún campo o la contraseña es muy corta. La
    duplicidad de nombre_usuario NO se valida aquí: se deja que la reporte el
    IntegrityError del commit, igual que ya hace num_licencia.
    """
    nombre_usuario = request.form.get("nombre_usuario", "").strip()
    contrasena = request.form.get("contrasena", "")
    email = request.form.get("email", "").strip()

    if not nombre_usuario:
        flash("El nombre de usuario es obligatorio.", "error")
        return None
    if len(contrasena) < 8:
        flash("La contraseña temporal debe tener al menos 8 caracteres.", "error")
        return None
    if not email:
        flash("El correo electrónico es obligatorio para crear la cuenta.", "error")
        return None

    return {"nombre_usuario": nombre_usuario, "contrasena": contrasena, "email": email}


def _datos_cuenta_mecanico_desde_formulario():
    """Lee y valida los campos del alta de una cuenta de mecánico (RF-1.1):
    nombre, nombre de usuario, contraseña temporal y correo, todos
    obligatorios (no hay tabla de perfil de mecánico: la cuenta Usuario es
    todo lo que existe para este rol).

    Retorna un dict con nombre/nombre_usuario/contrasena/email, o None (y ya
    hizo flash del error) si falta algún campo o la contraseña es muy corta.
    La duplicidad de nombre_usuario se deja al IntegrityError del commit,
    igual que en el alta de conductor.
    """
    nombre = request.form.get("nombre", "").strip()
    nombre_usuario = request.form.get("nombre_usuario", "").strip()
    contrasena = request.form.get("contrasena", "")
    email = request.form.get("email", "").strip()

    if not nombre:
        flash("El nombre es obligatorio.", "error")
        return None
    if not nombre_usuario:
        flash("El nombre de usuario es obligatorio.", "error")
        return None
    if len(contrasena) < 8:
        flash("La contraseña temporal debe tener al menos 8 caracteres.", "error")
        return None
    if not email:
        flash("El correo electrónico es obligatorio.", "error")
        return None

    return {
        "nombre": nombre,
        "nombre_usuario": nombre_usuario,
        "contrasena": contrasena,
        "email": email,
    }


def _datos_mecanico_editar_desde_formulario():
    """Lee y valida los campos editables de una cuenta de mecánico (RF-1.4):
    nombre, nombre de usuario y correo, todos obligatorios (mismo criterio
    que el alta). El rol NUNCA se lee del formulario aquí: esta vista no
    debe poder convertir un mecánico en admin.

    Retorna un dict con nombre/nombre_usuario/email, o None (y ya hizo flash
    del error) si falta algún campo. La duplicidad de nombre_usuario se deja
    al IntegrityError del commit, igual que en el resto del módulo.
    """
    nombre = request.form.get("nombre", "").strip()
    nombre_usuario = request.form.get("nombre_usuario", "").strip()
    email = request.form.get("email", "").strip()

    if not nombre:
        flash("El nombre es obligatorio.", "error")
        return None
    if not nombre_usuario:
        flash("El nombre de usuario es obligatorio.", "error")
        return None
    if not email:
        flash("El correo electrónico es obligatorio.", "error")
        return None

    return {"nombre": nombre, "nombre_usuario": nombre_usuario, "email": email}


def _datos_vehiculo_desde_formulario():
    """Lee y valida los campos comunes del formulario de vehículo.

    Retorna un dict listo para crear/actualizar un Vehiculo (sin 'estado', que
    se maneja aparte), o None (y ya hizo flash del error) si el año no es válido.

    Usada tanto por vehiculos_nuevo() como por vehiculos_editar() (alta y
    edición comparten esta misma función), así que normalizar aquí basta para
    cubrir ambos casos.

    placas se normaliza a mayúsculas (además del strip): el UNIQUE de BD sobre
    esa columna es sensible a mayúsculas/minúsculas, así que sin esto
    'hgo-1206' y 'HGO-1206' no chocarían entre sí y se colaría un vehículo con
    la misma placa física duplicada. vehiculos_placa_disponible() ya comparaba
    con func.upper() para el aviso en tiempo real; ahora el guardado real
    queda consistente con esa misma normalización.
    """
    anio_texto = request.form.get("anio", "")
    try:
        anio = int(anio_texto)
    except ValueError:
        flash("El año del vehículo no es válido.", "error")
        return None

    return {
        "placas": request.form.get("placas", "").strip().upper(),
        "num_unidad": request.form.get("num_unidad", "").strip(),
        "marca": request.form.get("marca", "").strip(),
        "modelo": request.form.get("modelo", "").strip(),
        "anio": anio,
        "num_serie": request.form.get("num_serie", "").strip(),
    }


def _vehiculo_para_formulario(vehiculo):
    """Convierte un Vehiculo de BD a dict apto para prellenar el formulario HTML."""
    return {
        "placas": vehiculo.placas,
        "num_unidad": vehiculo.num_unidad,
        "marca": vehiculo.marca,
        "modelo": vehiculo.modelo,
        "anio": vehiculo.anio,
        "num_serie": vehiculo.num_serie,
        "estado": vehiculo.estado,
    }


def _conteos_dashboard():
    """Los conteos de tarjetas del panel de administrador, compartidos por
    dashboard() y dashboard_estado() para que ambos siempre reporten
    exactamente lo mismo.

    La tarjeta "Emergencias" se compone de dos conteos separados en vez de
    uno solo: 'emergencia_mecanica' (alertas 'asistencia_mecanica' sin
    atender) y 'emergencia_trafico' (alertas 'incidencia_trafico' sin
    atender), con el mismo patrón para ambas. Antes incluía también
    'emergencia_seguridad' (viajes en estado 'emergencia'), pero ese conteo
    se quitó: el disparo manual del aviso silencioso (blueprint
    'emergencia', /emergencia/activar) ya no está registrado en
    app/__init__.py, así que ningún camino de código actual deja un Viaje en
    estado 'emergencia' -ese conteo siempre daría 0 y el desglose mentiría
    sobre por qué la tarjeta está en rojo. El estado 'emergencia' en sí
    sigue siendo válido en el modelo (CheckConstraint de Viaje) y se sigue
    consultando en otras vistas (viajes lista, cierre de viajes por admin,
    dashboard del conductor) por si queda algún viaje viejo en ese estado,
    pero ya no alimenta esta tarjeta.
    """
    return {
        "activo": Viaje.query.filter_by(estado="activo").count(),
        "alerta": Viaje.query.filter_by(estado="alerta").count(),
        "emergencia_mecanica": Alerta.query.filter_by(
            tipo="asistencia_mecanica", atendida=False
        ).count(),
        "emergencia_trafico": Alerta.query.filter_by(
            tipo="incidencia_trafico", atendida=False
        ).count(),
        "licencia_vencida": Conductor.query.filter(
            Conductor.fecha_vencimiento_lic < date.today()
        ).count(),
    }


def _texto_desglose_emergencia(conteos):
    """Texto tipo '1 mecánica · 2 de tráfico' para la tarjeta "Emergencias"
    del dashboard (ver _conteos_dashboard).
    """
    mecanica = conteos["emergencia_mecanica"]
    trafico = conteos["emergencia_trafico"]
    return (
        f"{mecanica} mecánica{'s' if mecanica != 1 else ''} · "
        f"{trafico} de tráfico"
    )


@admin.route("/perfil", methods=["GET", "POST"])
@admin_required
def perfil():
    """Pantalla "Mi perfil" del administrador con sesión iniciada: edición de
    correo/teléfono y cambio de contraseña propia (RF-1). La lógica es
    compartida con mecanico.perfil (ver app/controllers/perfil.py).
    """
    return procesar_perfil("admin/perfil.html")


@admin.route("/dashboard")
@admin_required
def dashboard():
    """Panel de administrador: 4 tarjetas de resumen, la misma lista de
    'Viajes en curso' que usa Viajes activos (ver _viajes_en_curso) y las
    alertas más recientes de cualquier tipo, atendidas o no.
    """
    conteos = _conteos_dashboard()

    viajes_en_curso = _viajes_en_curso()

    alertas_recientes = (
        Alerta.query.options(joinedload(Alerta.conductor))
        .order_by(Alerta.generada_en.desc())
        .limit(5)
        .all()
    )
    # Las 5 siguen siendo las más recientes de cualquier tipo (atendidas o
    # no); dentro de esas 5 se reordena para que las NO atendidas siempre
    # aparezcan primero, aunque sean un poco más viejas -lo urgente no debe
    # quedar enterrado debajo de algo ya resuelto. sort() es estable, así que
    # el orden por fecha dentro de cada grupo (atendida/no atendida) se
    # conserva.
    alertas_recientes.sort(key=lambda alerta: alerta.atendida)

    return render_template(
        "admin/dashboard.html",
        conteos=conteos,
        desglose_emergencia=_texto_desglose_emergencia(conteos),
        viajes_en_curso=viajes_en_curso,
        alertas_recientes=alertas_recientes,
    )


@admin.route("/dashboard/estado")
@admin_required
def dashboard_estado():
    """JSON con los mismos datos que dashboard() (4 conteos + ids/timestamps
    de las últimas 5 alertas), para que el panel detecte cambios en segundo
    plano (polling cada 30s) sin recargar la página completa.
    """
    conteos = _conteos_dashboard()

    alertas_recientes = Alerta.query.order_by(Alerta.generada_en.desc()).limit(5).all()

    return jsonify(
        conteos=conteos,
        alertas=[
            {
                "id": alerta.id_alerta,
                "atendida": alerta.atendida,
                "generada_en": alerta.generada_en.isoformat(),
            }
            for alerta in alertas_recientes
        ],
    )


def _viajes_programados():
    """Viajes en estado 'programado' (ruta asignada por el admin, todavía sin
    iniciar por el conductor), con conductor y vehículo precargados.

    Se usa tanto en vehiculos_lista como en conductores_lista para el
    indicador "Reservado": mientras un Viaje sigue 'programado', el
    conductor y el vehículo que le asignó el admin ya están comprometidos
    con esa ruta, aunque Vehiculo.estado siga en 'disponible' -su CHECK
    constraint no tiene un valor 'reservado' y no se modifica aquí- y la
    cuenta del conductor no tenga ninguna marca propia. Por eso se deriva en
    Python cruzando contra Viaje en vez de agregar una columna nueva, igual
    que ya hace mecanico.vehiculos_lista con la alerta mecánica pendiente.
    """
    return (
        Viaje.query.filter_by(estado="programado")
        .options(joinedload(Viaje.conductor), joinedload(Viaje.vehiculo))
        .all()
    )


_ETIQUETA_GRUPO_CONDUCTOR = {
    0: "Licencia vigente",
    1: "Licencia próxima a vencer",
    2: "Licencia vencida",
    3: "Cuenta inactiva o sin vincular",
}


def _prioridad_orden_conductor(conductor):
    """Prioridad de orden por default de conductores_lista (orden lógico de
    UX, no un RF): primero las cuentas activas, agrupadas entre sí por
    vigencia de licencia (vigente -> próxima a vencer -> vencida); al final,
    en un solo grupo sin importar la licencia, las cuentas inactivas o sin
    cuenta vinculada -ninguna de las dos puede iniciar sesión de todas
    formas, así que no vale la pena subdividirlas por licencia.

    Reutiliza licencia_vigente()/licencia_proxima_a_vencer() del modelo
    (no son columnas, son métodos calculados) en vez de duplicar esa lógica
    aquí: por eso el orden se resuelve en Python después de traer la lista,
    no con un ORDER BY de SQL.
    """
    cuenta = conductor.usuarios[0] if conductor.usuarios else None
    if cuenta is None or not cuenta.activo:
        return 3
    if not conductor.licencia_vigente():
        return 2
    if conductor.licencia_proxima_a_vencer():
        return 1
    return 0


def _clave_orden_conductor(conductor):
    return (_prioridad_orden_conductor(conductor), _normalizar_busqueda(conductor.nombre))


@admin.route("/conductores")
@admin_required
def conductores_lista():
    """Lista los conductores con el estado de vigencia de su licencia (RF-2.3),
    con 4 tarjetas de resumen, un buscador simple por nombre (?buscar=), un
    filtro rápido por estado de cuenta (?estado_cuenta=activos|inactivos) y
    otro por vigencia de licencia (?licencia=vencida, usado por la tarjeta
    "Licencia vencida" del dashboard), combinables entre sí.

    Los conteos se calculan sobre TODOS los conductores (no sobre el
    resultado ya filtrado), para que las tarjetas reflejen siempre el estado
    real de la flota completa sin importar los filtros activos -igual que
    en el mockup, donde los números no cambian al buscar.

    El orden por default agrupa por prioridad operativa (ver
    _prioridad_orden_conductor) y alfabético por nombre dentro de cada
    grupo; se resuelve siempre en Python, incluso cuando hay filtros
    activos, para que el orden se sienta consistente sea cual sea el
    resultado visible.
    """
    todos = Conductor.query.order_by(Conductor.nombre).all()

    conteos = {
        "total": len(todos),
        "vigentes": sum(
            1 for c in todos if c.licencia_vigente() and not c.licencia_proxima_a_vencer()
        ),
        "proximas": sum(1 for c in todos if c.licencia_proxima_a_vencer()),
        "vencidas": sum(1 for c in todos if not c.licencia_vigente()),
    }

    buscar = request.args.get("buscar", "").strip()
    conductores = todos
    if buscar:
        buscar_normalizado = _normalizar_busqueda(buscar)
        conductores = [
            c for c in conductores if buscar_normalizado in _normalizar_busqueda(c.nombre)
        ]

    estado_cuenta = request.args.get("estado_cuenta", "").strip()
    if estado_cuenta == "activos":
        conductores = [c for c in conductores if c.usuarios and c.usuarios[0].activo]
    elif estado_cuenta == "inactivos":
        conductores = [c for c in conductores if not (c.usuarios and c.usuarios[0].activo)]

    licencia_filtro = request.args.get("licencia", "").strip()
    if licencia_filtro == "vencida":
        conductores = [c for c in conductores if not c.licencia_vigente()]

    conductores = sorted(conductores, key=_clave_orden_conductor)
    grupo_conductor = {
        conductor.id_conductor: _ETIQUETA_GRUPO_CONDUCTOR[_prioridad_orden_conductor(conductor)]
        for conductor in conductores
    }

    # Ruta programada pendiente por conductor (ver _viajes_programados): un
    # conductor por ruta programada alcanza para el indicador visual, no
    # hace falta guardar todas si por alguna razón tuviera más de una.
    reserva_por_conductor = {}
    for viaje in _viajes_programados():
        reserva_por_conductor.setdefault(viaje.id_conductor, viaje)

    return render_template(
        "admin/conductores/lista.html",
        conductores=conductores,
        conteos=conteos,
        buscar=buscar,
        estado_cuenta=estado_cuenta,
        licencia_filtro=licencia_filtro,
        grupo_conductor=grupo_conductor,
        reserva_por_conductor=reserva_por_conductor,
    )


@admin.route("/conductores/nuevo", methods=["GET", "POST"])
@admin_required
def conductores_nuevo():
    """Muestra el formulario de alta y crea el Conductor junto con su cuenta
    de acceso Usuario, en la misma transacción (RF-2.1 + RF-1.1): si algo
    falla no debe quedar un Conductor huérfano sin cuenta ni una cuenta sin
    Conductor.
    """
    if request.method == "POST":
        datos_conductor = _datos_conductor_desde_formulario()
        if datos_conductor is None:
            return render_template(
                "admin/conductores/formulario.html", conductor=request.form, id_conductor=None
            )

        datos_cuenta = _datos_cuenta_nueva_desde_formulario()
        if datos_cuenta is None:
            return render_template(
                "admin/conductores/formulario.html", conductor=request.form, id_conductor=None
            )

        conductor = Conductor(**datos_conductor)
        usuario = Usuario(
            nombre=datos_conductor["nombre"],
            nombre_usuario=datos_cuenta["nombre_usuario"],
            email=datos_cuenta["email"],
            rol="conductor",
            conductor=conductor,
        )
        usuario.set_password(datos_cuenta["contrasena"])

        try:
            db.session.add(conductor)
            db.session.add(usuario)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            if Conductor.query.filter_by(num_licencia=datos_conductor["num_licencia"]).first():
                flash(
                    f"Ya existe un conductor con el número de licencia '{datos_conductor['num_licencia']}'.",
                    "error",
                )
            elif Usuario.query.filter_by(nombre_usuario=datos_cuenta["nombre_usuario"]).first():
                flash(
                    f"Ya existe una cuenta con el nombre de usuario '{datos_cuenta['nombre_usuario']}'.",
                    "error",
                )
            else:
                flash("No se pudo registrar el conductor. Intenta de nuevo.", "error")
            return render_template(
                "admin/conductores/formulario.html", conductor=request.form, id_conductor=None
            )

        bitacora = Bitacora(
            id_usuario=current_user.id_usuario,
            accion="alta_cuenta_conductor",
            descripcion=(
                f"Alta de conductor '{conductor.nombre}' con cuenta de usuario "
                f"'{usuario.nombre_usuario}' por {current_user.nombre}."
            ),
            tabla_afectada="usuarios",
            registro_id=usuario.id_usuario,
        )
        db.session.add(bitacora)
        db.session.commit()

        flash(
            f"Conductor '{conductor.nombre}' y su cuenta de acceso fueron registrados correctamente.",
            "success",
        )
        return redirect(url_for("admin.conductores_lista"))

    return render_template("admin/conductores/formulario.html", conductor=None, id_conductor=None)


@admin.route("/conductores/<int:id_conductor>/editar", methods=["GET", "POST"])
@admin_required
def conductores_editar(id_conductor):
    """Muestra el formulario prellenado y actualiza los datos del conductor (RF-2.1)."""
    conductor = Conductor.query.get_or_404(id_conductor)

    if request.method == "POST":
        datos = _datos_conductor_desde_formulario()
        if datos is None:
            return render_template(
                "admin/conductores/formulario.html",
                conductor=request.form,
                id_conductor=id_conductor,
            )

        # La nueva contraseña es opcional: un conductor sin correo registrado
        # no puede usar "olvidé mi contraseña", así que esta es la única vía
        # para ayudarlo si la olvida (mismo patrón que mecanicos_editar).
        # Vacío = no se toca la actual.
        nueva_contrasena = request.form.get("nueva_contrasena", "")
        if nueva_contrasena and len(nueva_contrasena) < 8:
            flash("La nueva contraseña debe tener al menos 8 caracteres.", "error")
            return render_template(
                "admin/conductores/formulario.html",
                conductor=request.form,
                id_conductor=id_conductor,
            )

        for campo, valor in datos.items():
            setattr(conductor, campo, valor)

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash(
                f"Ya existe un conductor con el número de licencia '{datos['num_licencia']}'.",
                "error",
            )
            return render_template(
                "admin/conductores/formulario.html",
                conductor=request.form,
                id_conductor=id_conductor,
            )

        _guardar_email_cuenta(conductor, request.form.get("email", "").strip())

        contrasena_cambiada = False
        if nueva_contrasena:
            contrasena_cambiada = _guardar_contrasena_cuenta(conductor, nueva_contrasena)

        bitacora = Bitacora(
            id_usuario=current_user.id_usuario,
            accion="edicion_conductor",
            descripcion=(
                f"Conductor '{conductor.nombre}' editado por {current_user.nombre}"
                + (" (incluye cambio de contraseña)." if contrasena_cambiada else ".")
            ),
            tabla_afectada="conductores",
            registro_id=conductor.id_conductor,
        )
        db.session.add(bitacora)
        db.session.commit()

        flash(f"Conductor '{conductor.nombre}' actualizado correctamente.", "success")
        return redirect(url_for("admin.conductores_lista"))

    return render_template(
        "admin/conductores/formulario.html",
        conductor=_conductor_para_formulario(conductor),
        id_conductor=id_conductor,
    )


@admin.route("/conductores/<int:id_conductor>/crear-cuenta", methods=["GET", "POST"])
@admin_required
def conductores_crear_cuenta(id_conductor):
    """Crea la cuenta de acceso Usuario para un Conductor que ya existía sin
    una (badge "Sin cuenta" en conductores_lista, RF-1.1 aplicado
    retroactivamente). Distinto de conductores_nuevo(), donde Conductor y
    Usuario se crean juntos en la misma transacción: aquí el Conductor ya
    existe de antes y solo falta darle acceso.

    Solo aplica si el conductor todavía NO tiene ninguna cuenta vinculada:
    si ya tiene una, esta no es la ruta para tocarla (eso es
    conductores_editar + _guardar_email_cuenta), así que responde 404.
    """
    conductor = Conductor.query.get_or_404(id_conductor)

    if Usuario.query.filter_by(id_conductor=conductor.id_conductor).first() is not None:
        abort(404)

    if request.method == "POST":
        datos_cuenta = _datos_cuenta_nueva_desde_formulario()
        if datos_cuenta is None:
            return render_template(
                "admin/conductores/crear_cuenta.html", conductor=conductor, formulario=request.form
            )

        usuario = Usuario(
            nombre=conductor.nombre,
            nombre_usuario=datos_cuenta["nombre_usuario"],
            email=datos_cuenta["email"],
            rol="conductor",
            id_conductor=conductor.id_conductor,
        )
        usuario.set_password(datos_cuenta["contrasena"])

        try:
            db.session.add(usuario)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash(
                f"Ya existe una cuenta con el nombre de usuario '{datos_cuenta['nombre_usuario']}'.",
                "error",
            )
            return render_template(
                "admin/conductores/crear_cuenta.html", conductor=conductor, formulario=request.form
            )

        bitacora = Bitacora(
            id_usuario=current_user.id_usuario,
            accion="alta_cuenta_conductor_retroactiva",
            descripcion=(
                f"Cuenta de acceso creada de forma retroactiva para el conductor "
                f"'{conductor.nombre}' (usuario '{usuario.nombre_usuario}') por "
                f"{current_user.nombre}."
            ),
            tabla_afectada="usuarios",
            registro_id=usuario.id_usuario,
        )
        db.session.add(bitacora)
        db.session.commit()

        flash(f"Cuenta de acceso creada correctamente para '{conductor.nombre}'.", "success")
        return redirect(url_for("admin.conductores_lista"))

    return render_template("admin/conductores/crear_cuenta.html", conductor=conductor, formulario=None)


@admin.route("/mecanicos")
@admin_required
def mecanicos_lista():
    """Lista todas las cuentas de mecánico (RF-1.4: el admin gestiona cuentas).

    No existe una tabla de perfil de mecánico -a diferencia de Conductor-,
    así que esta lista consulta directamente Usuario filtrando por rol.
    """
    mecanicos = Usuario.query.filter_by(rol="mecanico").order_by(Usuario.nombre).all()
    return render_template("admin/mecanicos/lista.html", mecanicos=mecanicos)


@admin.route("/mecanicos/nuevo", methods=["GET", "POST"])
@admin_required
def mecanicos_nuevo():
    """Muestra el formulario de alta y crea la cuenta Usuario con rol
    mecánico (RF-1.1). A diferencia de conductores, aquí no hay perfil
    aparte que crear: la cuenta Usuario es todo lo que existe para este rol.
    """
    if request.method == "POST":
        datos_cuenta = _datos_cuenta_mecanico_desde_formulario()
        if datos_cuenta is None:
            return render_template("admin/mecanicos/formulario.html", mecanico=request.form)

        usuario = Usuario(
            nombre=datos_cuenta["nombre"],
            nombre_usuario=datos_cuenta["nombre_usuario"],
            email=datos_cuenta["email"],
            rol="mecanico",
        )
        usuario.set_password(datos_cuenta["contrasena"])

        try:
            db.session.add(usuario)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash(
                f"Ya existe una cuenta con el nombre de usuario '{datos_cuenta['nombre_usuario']}'.",
                "error",
            )
            return render_template("admin/mecanicos/formulario.html", mecanico=request.form)

        bitacora = Bitacora(
            id_usuario=current_user.id_usuario,
            accion="alta_cuenta_mecanico",
            descripcion=(
                f"Alta de cuenta de mecánico '{usuario.nombre}' (usuario "
                f"'{usuario.nombre_usuario}') por {current_user.nombre}."
            ),
            tabla_afectada="usuarios",
            registro_id=usuario.id_usuario,
        )
        db.session.add(bitacora)
        db.session.commit()

        flash(f"Cuenta de mecánico '{usuario.nombre}' registrada correctamente.", "success")
        return redirect(url_for("admin.mecanicos_lista"))

    return render_template("admin/mecanicos/formulario.html", mecanico=None, id_usuario=None)


@admin.route("/mecanicos/<int:id_usuario>/editar", methods=["GET", "POST"])
@admin_required
def mecanicos_editar(id_usuario):
    """Muestra el formulario prellenado y actualiza nombre/usuario/correo de
    una cuenta de mecánico (RF-1.4), con cambio de contraseña opcional.

    Se filtra por rol='mecanico' al buscar la cuenta (no solo por
    id_usuario) para que esta vista nunca pueda editar, ni por error de URL,
    una cuenta de admin o de conductor. El rol tampoco se toca al guardar:
    este formulario no ofrece ese campo.
    """
    mecanico = Usuario.query.filter_by(id_usuario=id_usuario, rol="mecanico").first_or_404()

    if request.method == "POST":
        datos = _datos_mecanico_editar_desde_formulario()
        if datos is None:
            return render_template(
                "admin/mecanicos/formulario.html",
                mecanico=request.form,
                id_usuario=id_usuario,
            )

        # La nueva contraseña es opcional: un mecánico sin correo registrado
        # no puede usar "olvidé mi contraseña", así que esta es la única vía
        # para ayudarlo si la olvida. Vacío = no se toca la actual.
        nueva_contrasena = request.form.get("nueva_contrasena", "")
        if nueva_contrasena and len(nueva_contrasena) < 8:
            flash("La nueva contraseña debe tener al menos 8 caracteres.", "error")
            return render_template(
                "admin/mecanicos/formulario.html",
                mecanico=request.form,
                id_usuario=id_usuario,
            )

        for campo, valor in datos.items():
            setattr(mecanico, campo, valor)

        if nueva_contrasena:
            mecanico.set_password(nueva_contrasena)

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash(
                f"Ya existe una cuenta con el nombre de usuario '{datos['nombre_usuario']}'.",
                "error",
            )
            return render_template(
                "admin/mecanicos/formulario.html",
                mecanico=request.form,
                id_usuario=id_usuario,
            )

        bitacora = Bitacora(
            id_usuario=current_user.id_usuario,
            accion="edicion_cuenta_mecanico",
            descripcion=(
                f"Cuenta de mecánico '{mecanico.nombre}' (usuario "
                f"'{mecanico.nombre_usuario}') editada por {current_user.nombre}"
                + (" (incluye cambio de contraseña)." if nueva_contrasena else ".")
            ),
            tabla_afectada="usuarios",
            registro_id=mecanico.id_usuario,
        )
        db.session.add(bitacora)
        db.session.commit()

        flash(f"Cuenta de mecánico '{mecanico.nombre}' actualizada correctamente.", "success")
        return redirect(url_for("admin.mecanicos_lista"))

    return render_template(
        "admin/mecanicos/formulario.html",
        mecanico={
            "nombre": mecanico.nombre,
            "nombre_usuario": mecanico.nombre_usuario,
            "email": mecanico.email or "",
        },
        id_usuario=id_usuario,
    )


_DESTINO_POR_ROL_CUENTA = {
    "mecanico": "admin.mecanicos_lista",
    "conductor": "admin.conductores_lista",
}


@admin.route("/cuentas/<int:id_usuario>/estado", methods=["POST"])
@admin_required
def cuentas_cambiar_estado(id_usuario):
    """Activa o desactiva una cuenta de usuario (RF-1.4), usada por las listas
    de Mecánicos y Conductores. Comparte una sola ruta porque el toggle es
    idéntico para ambos roles; solo cambia a dónde se redirige después.

    Dos protecciones para no dejar el sistema sin nadie que lo administre:
    un admin no puede desactivar su propia cuenta, y no puede desactivar al
    último administrador activo que quede (aunque no sea él mismo).

    Desactivar la cuenta de un conductor NUNCA toca sus Viajes/Alertas
    pasados ni los oculta de ningún lado (Historial, Viajes activos,
    Bitácora, etc.): esos filtran por estado del Viaje, no por si la cuenta
    del conductor sigue activa. `usuario.activo` solo lo lee is_active()
    (bloquea el login, ver auth.login) y esta misma vista.
    """
    usuario = Usuario.query.get_or_404(id_usuario)
    destino = _DESTINO_POR_ROL_CUENTA.get(usuario.rol, "admin.dashboard")

    if usuario.id_usuario == current_user.id_usuario:
        flash("No puedes desactivar tu propia cuenta.", "error")
        return redirect(url_for(destino))

    if usuario.rol == "admin" and usuario.activo:
        admins_activos = Usuario.query.filter_by(rol="admin", activo=True).count()
        if admins_activos <= 1:
            flash("No puedes desactivar al único administrador activo del sistema.", "error")
            return redirect(url_for(destino))

    # Decisión de diseño: NO se bloquea desactivar a un conductor con un
    # viaje 'programado'/'activo'/'alerta'/'emergencia' sin cerrar -- el
    # admin puede necesitar quitarle el acceso de inmediato (ej. renuncia a
    # medio turno) sin que eso dependa de resolver primero ese viaje, que de
    # todas formas se sigue gestionando igual desde Viajes activos / Rutas
    # programadas (cerrar forzado, cancelar programada) sin importar si el
    # conductor puede volver a iniciar sesión. Además, una sesión YA
    # iniciada no se cierra sola al desactivar la cuenta (Flask-Login no
    # revalida is_active en cada request), así que un conductor a mitad de
    # viaje normalmente puede seguir confirmando su llegada aunque se le
    # desactive la cuenta en ese momento. Se permite, pero se avisa
    # explícitamente para que el admin no lo pierda de vista.
    viaje_sin_cerrar = None
    if usuario.rol == "conductor" and usuario.activo and usuario.id_conductor is not None:
        viaje_sin_cerrar = Viaje.query.filter(
            Viaje.id_conductor == usuario.id_conductor,
            Viaje.estado.in_(["programado", "activo", "alerta", "emergencia"]),
        ).first()

    estado_anterior = "activa" if usuario.activo else "inactiva"
    usuario.activo = not usuario.activo
    estado_nuevo = "activa" if usuario.activo else "inactiva"

    bitacora = Bitacora(
        id_usuario=current_user.id_usuario,
        accion="cambio_estado_cuenta",
        descripcion=(
            f"Cuenta '{usuario.nombre_usuario}' cambiada de {estado_anterior} a "
            f"{estado_nuevo} por {current_user.nombre}."
        ),
        tabla_afectada="usuarios",
        registro_id=usuario.id_usuario,
    )
    db.session.add(bitacora)
    db.session.commit()

    flash(f"Cuenta '{usuario.nombre_usuario}' ahora está {estado_nuevo}.", "success")

    if viaje_sin_cerrar is not None:
        flash(
            f"Nota: {usuario.nombre} tenía un viaje '{viaje_sin_cerrar.estado}' sin cerrar "
            f"(#{viaje_sin_cerrar.id_viaje}) al desactivarlo. Eso NO se canceló ni se cerró "
            "solo -- únicamente se le bloqueó el inicio de sesión. Revisa Viajes activos o "
            "Rutas programadas si hace falta cerrarlo o cancelarlo.",
            "warning",
        )
    return redirect(url_for(destino))


_PRIORIDAD_ESTADO_VEHICULO = {"disponible": 0, "en_ruta": 1, "en_taller": 2}
_ETIQUETA_GRUPO_VEHICULO = {
    "disponible": "Disponibles",
    "en_ruta": "En ruta",
    "en_taller": "En taller",
}


def _clave_orden_vehiculo(vehiculo):
    """Orden por default de vehiculos_lista (orden lógico de UX, no un RF):
    Disponibles primero -son los que el admin necesita ver para asignar una
    ruta nueva-, luego En ruta, luego En taller; alfabético por placa dentro
    de cada grupo. `estado` sí es una columna directa, así que a diferencia
    de conductores_lista esto podría resolverse con SQL, pero se deja en
    Python para no mezclar dos técnicas de orden distintas entre ambas
    listas -la función ya trae `todos` completo a Python de todas formas
    para calcular los conteos.
    """
    return (
        _PRIORIDAD_ESTADO_VEHICULO.get(vehiculo.estado, 99),
        _normalizar_busqueda(vehiculo.placas),
    )


@admin.route("/vehiculos")
@admin_required
def vehiculos_lista():
    """Lista los vehículos con su estado actual (RF-2.3), con 5 tarjetas de
    resumen y un buscador simple por placa (?buscar=).

    Igual que en conductores_lista, los conteos se calculan sobre TODOS los
    vehículos, no sobre el resultado ya filtrado por ?buscar=. El orden por
    default agrupa por estado operativo (ver _clave_orden_vehiculo) -un
    vehículo reservado sigue con estado='disponible' en la BD (ver
    _viajes_programados), así que se queda en el mismo grupo "Disponibles";
    no se creó un grupo aparte porque el badge en la fila ya lo distingue.

    Decisión sobre las tarjetas: un vehículo reservado NO cuenta como
    "Disponible" (ya está comprometido con una ruta), así que se resta de
    ese conteo; en vez de perder ese número se agrega una tarjeta
    "Reservados" nueva, para no esconder cuántos vehículos están en esa
    situación.
    """
    todos = Vehiculo.query.order_by(Vehiculo.placas).all()

    # Vehículos con una ruta programada pendiente de iniciar (ver
    # _viajes_programados): aunque su columna `estado` siga en 'disponible',
    # ya están comprometidos y el admin no debería volver a asignarlos.
    reserva_por_vehiculo = {}
    for viaje in _viajes_programados():
        reserva_por_vehiculo.setdefault(viaje.id_vehiculo, viaje)

    conteos = {
        "total": len(todos),
        "disponibles": sum(
            1
            for v in todos
            if v.estado == "disponible" and v.id_vehiculo not in reserva_por_vehiculo
        ),
        "reservados": sum(
            1
            for v in todos
            if v.estado == "disponible" and v.id_vehiculo in reserva_por_vehiculo
        ),
        "en_ruta": sum(1 for v in todos if v.estado == "en_ruta"),
        "en_taller": sum(1 for v in todos if v.estado == "en_taller"),
    }

    buscar = request.args.get("buscar", "").strip()
    vehiculos = todos
    if buscar:
        buscar_normalizado = _normalizar_busqueda(buscar)
        vehiculos = [
            v for v in vehiculos if buscar_normalizado in _normalizar_busqueda(v.placas)
        ]

    vehiculos = sorted(vehiculos, key=_clave_orden_vehiculo)
    grupo_vehiculo = {
        vehiculo.id_vehiculo: _ETIQUETA_GRUPO_VEHICULO.get(vehiculo.estado, "Otro")
        for vehiculo in vehiculos
    }

    return render_template(
        "admin/vehiculos/lista.html",
        vehiculos=vehiculos,
        conteos=conteos,
        buscar=buscar,
        grupo_vehiculo=grupo_vehiculo,
        reserva_por_vehiculo=reserva_por_vehiculo,
    )


def _sugerir_num_unidad():
    """Sugiere el siguiente número de unidad consecutivo para prellenar el
    alta de un vehículo nuevo (RF-2.2): toma el num_unidad numérico más alto
    ya registrado y propone ese valor + 1, conservando el mismo relleno de
    ceros que tenía ese registro (ej. '01' -> '02'). Es solo una sugerencia
    prellenada -el admin puede cambiarla libremente-, no un valor bloqueado
    ni validado.

    num_unidad es texto libre (ver Vehiculo.num_unidad), así que puede haber
    registros viejos no puramente numéricos; esos se ignoran para este
    cálculo en vez de tronar. Si no hay ningún num_unidad numérico
    registrado todavía, sugiere '01'.
    """
    maximo_valor = None
    maximo_texto = None
    for (num_unidad,) in db.session.query(Vehiculo.num_unidad).all():
        texto = (num_unidad or "").strip()
        if not texto.isdigit():
            continue
        valor = int(texto)
        if maximo_valor is None or valor > maximo_valor:
            maximo_valor = valor
            maximo_texto = texto

    if maximo_valor is None:
        return "01"
    return str(maximo_valor + 1).zfill(len(maximo_texto))


@admin.route("/vehiculos/placa-disponible")
@admin_required
def vehiculos_placa_disponible():
    """Ayuda de UX para el formulario de vehículo (alta y edición): indica si
    ya existe un vehículo con esa placa, para avisar en tiempo real antes de
    enviar el formulario.

    Esto NO reemplaza la validación real: el IntegrityError + rollback al
    guardar (ver vehiculos_nuevo/vehiculos_editar) sigue siendo la única
    protección definitiva contra una placa duplicada, por si el JS está
    deshabilitado o hay una condición de carrera entre esta consulta y el
    guardado real.

    ?excluir=<id_vehiculo> excluye ese vehículo de la comparación, para que
    el formulario de edición no marque la placa actual del vehículo como
    duplicada de sí misma.
    """
    placa = request.args.get("placa", "").strip().upper()
    if not placa:
        return jsonify(disponible=True)

    id_excluir = request.args.get("excluir", type=int)
    query = Vehiculo.query.filter(db.func.upper(Vehiculo.placas) == placa)
    if id_excluir is not None:
        query = query.filter(Vehiculo.id_vehiculo != id_excluir)

    return jsonify(disponible=query.first() is None)


@admin.route("/vehiculos/nuevo", methods=["GET", "POST"])
@admin_required
def vehiculos_nuevo():
    """Muestra el formulario de alta y crea el Vehiculo al enviarlo (RF-2.2).

    El estado no se pide aquí: todo vehículo nuevo inicia en 'disponible'.
    """
    if request.method == "POST":
        datos = _datos_vehiculo_desde_formulario()
        if datos is None:
            return render_template(
                "admin/vehiculos/formulario.html", vehiculo=request.form, id_vehiculo=None
            )

        vehiculo = Vehiculo(**datos)
        try:
            db.session.add(vehiculo)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash(
                f"Ya existe un vehículo con las placas '{datos['placas']}'.",
                "error",
            )
            return render_template(
                "admin/vehiculos/formulario.html", vehiculo=request.form, id_vehiculo=None
            )

        flash(f"Vehículo '{vehiculo.placas}' registrado correctamente.", "success")
        return redirect(url_for("admin.vehiculos_lista"))

    return render_template(
        "admin/vehiculos/formulario.html",
        vehiculo={"num_unidad": _sugerir_num_unidad()},
        id_vehiculo=None,
    )


@admin.route("/vehiculos/<int:id_vehiculo>/editar", methods=["GET", "POST"])
@admin_required
def vehiculos_editar(id_vehiculo):
    """Muestra el formulario prellenado y actualiza los datos del vehículo (RF-2.2).

    El estado solo puede cambiarse aquí entre 'disponible' y 'en_taller'. Si el
    vehículo está actualmente 'en_ruta' (viaje activo, RF-5.5), el estado no es
    editable desde este formulario bajo ninguna circunstancia: cualquier POST que
    incluya el campo 'estado' se rechaza, sin importar su valor.
    """
    vehiculo = Vehiculo.query.get_or_404(id_vehiculo)

    if request.method == "POST":
        if vehiculo.estado == "en_ruta" and "estado" in request.form:
            flash(
                "No se puede modificar el estado de un vehículo en ruta; "
                "se actualizará automáticamente al finalizar el viaje activo.",
                "error",
            )
            return render_template(
                "admin/vehiculos/formulario.html",
                vehiculo=request.form,
                id_vehiculo=id_vehiculo,
            )

        datos = _datos_vehiculo_desde_formulario()
        if datos is None:
            return render_template(
                "admin/vehiculos/formulario.html",
                vehiculo=request.form,
                id_vehiculo=id_vehiculo,
            )

        if vehiculo.estado != "en_ruta":
            estado = request.form.get("estado", "")
            if estado not in ("disponible", "en_taller"):
                flash("El estado seleccionado no es válido.", "error")
                return render_template(
                    "admin/vehiculos/formulario.html",
                    vehiculo=request.form,
                    id_vehiculo=id_vehiculo,
                )
            datos["estado"] = estado

        for campo, valor in datos.items():
            setattr(vehiculo, campo, valor)

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash(
                f"Ya existe un vehículo con las placas '{datos['placas']}'.",
                "error",
            )
            return render_template(
                "admin/vehiculos/formulario.html",
                vehiculo=request.form,
                id_vehiculo=id_vehiculo,
            )

        flash(f"Vehículo '{vehiculo.placas}' actualizado correctamente.", "success")
        return redirect(url_for("admin.vehiculos_lista"))

    return render_template(
        "admin/vehiculos/formulario.html",
        vehiculo=_vehiculo_para_formulario(vehiculo),
        id_vehiculo=id_vehiculo,
    )


def _viajes_en_curso(estados=("activo", "alerta", "emergencia")):
    """Viajes 'activo', 'alerta' o 'emergencia' para monitoreo del administrador
    (RF-3/RF-4). Compartida por viajes_lista() (vista de gestión completa) y
    dashboard() (resumen 'Viajes en curso'), para que ambas pantallas
    muestren siempre exactamente el mismo conjunto de viajes.

    Las emergencias reales van siempre primero, sin importar su hora_salida:
    son lo más urgente y no deben poder quedar enterradas debajo de simples
    retrasos o viajes normales más recientes.

    estados (opcional): subconjunto de los tres a incluir -usado por
    viajes_lista para el filtro ?estado= que enlazan las tarjetas del
    dashboard ("Operación normal"/"Retraso detectado", ver ahí-. Por
    default incluye los tres, igual que antes; dashboard() sigue llamando
    esta función sin argumentos, así que su resumen "Viajes en curso" no
    cambia.
    """
    prioridad_estado = case(
        (Viaje.estado == "emergencia", 0),
        (Viaje.estado == "alerta", 1),
        else_=2,
    )
    return (
        Viaje.query.filter(Viaje.estado.in_(estados))
        .order_by(prioridad_estado, Viaje.hora_salida)
        .all()
    )


def _detalle_viaje(viaje):
    """Arma los datos derivados para el panel de detalle de un viaje (RF-3):
    los minutos de retraso (calculados en el momento, no un campo de BD) y
    la línea de tiempo de eventos que sí se pueden derivar de datos reales.
    """
    ahora = datetime.now()

    retraso_minutos = None
    if viaje.hora_llegada is None and ahora > viaje.eta:
        retraso_minutos = int((ahora - viaje.eta).total_seconds() // 60)

    texto_por_tipo_alerta = {
        "retraso": "Alerta generada automáticamente",
        "panico": "Aviso de emergencia enviado",
        "licencia_vencida": "Alerta de licencia vencida generada",
    }

    eventos = [{"hora": viaje.hora_salida, "texto": "Viaje iniciado"}]

    # El Figma muestra "check-in con licencia vigente" como un paso separado
    # de la línea de tiempo, pero en el modelo real esa verificación ocurre
    # dentro de la misma transacción de inicio de viaje (RF-6.2, ver
    # conductor.py:viajes_iniciar) y no se guarda como un timestamp propio:
    # no hay un segundo dato real que mostrar. Se omite en vez de inventar
    # una hora que no existe; "Viaje iniciado" (hora_salida) ya representa
    # ese instante, licencia verificada incluida.

    alertas_del_viaje = (
        Alerta.query.filter_by(id_viaje=viaje.id_viaje).order_by(Alerta.generada_en).all()
    )
    for alerta in alertas_del_viaje:
        eventos.append(
            {
                "hora": alerta.generada_en,
                "texto": texto_por_tipo_alerta.get(alerta.tipo, "Alerta generada"),
            }
        )

    if viaje.hora_llegada is not None:
        eventos.append({"hora": viaje.hora_llegada, "texto": "Viaje cerrado"})

    eventos.sort(key=lambda evento: evento["hora"])

    return {"ahora": ahora, "retraso_minutos": retraso_minutos, "eventos": eventos}


_ESTADOS_FILTRABLES_VIAJES_LISTA = {"activo": ("activo",), "alerta": ("alerta",)}


@admin.route("/viajes")
@admin_required
def viajes_lista():
    """Vista de dos columnas: lista de viajes en curso a la izquierda y el
    detalle del seleccionado a la derecha (?id_viaje=<id>; si falta o no
    corresponde a un viaje en curso, se usa el primero de la lista, RF-3).

    ?estado=activo|alerta (opcional) reduce la lista a un solo estado -usado
    por las tarjetas "Operación normal"/"Retraso detectado" del dashboard,
    ver ahí-; cualquier otro valor (incluida su ausencia) conserva el
    comportamiento default: los tres estados en curso mezclados.
    """
    estado_filtro = request.args.get("estado", "")
    estados = _ESTADOS_FILTRABLES_VIAJES_LISTA.get(
        estado_filtro, ("activo", "alerta", "emergencia")
    )
    if estado_filtro not in _ESTADOS_FILTRABLES_VIAJES_LISTA:
        estado_filtro = ""
    viajes = _viajes_en_curso(estados=estados)

    id_viaje_param = request.args.get("id_viaje", type=int)
    viaje_seleccionado = None
    if id_viaje_param is not None:
        viaje_seleccionado = next(
            (v for v in viajes if v.id_viaje == id_viaje_param), None
        )
    if viaje_seleccionado is None and viajes:
        viaje_seleccionado = viajes[0]

    detalle = _detalle_viaje(viaje_seleccionado) if viaje_seleccionado else None

    return render_template(
        "admin/viajes/lista.html",
        viajes=viajes,
        viaje_seleccionado=viaje_seleccionado,
        detalle=detalle,
        estado_filtro=estado_filtro,
    )


@admin.route("/viajes/<int:id_viaje>/confirmar-llegada-manual", methods=["POST"])
@admin_required
def viajes_confirmar_llegada_manual(id_viaje):
    """Confirma la llegada de un viaje en nombre del conductor (mismo efecto
    que conductor.viajes_confirmar_llegada). Pensada para cuando el
    conductor no puede confirmar él mismo pero el viaje sí terminó bien."""
    viaje = Viaje.query.get_or_404(id_viaje)

    if viaje.estado not in VIAJE_ESTADOS_CERRABLES_POR_ADMIN:
        flash("Ese viaje ya no está en curso.", "error")
        return redirect(url_for("admin.viajes_lista"))

    estado_anterior = viaje.estado
    viaje.hora_llegada = datetime.now()
    viaje.estado = "completado"
    viaje.vehiculo.estado = "disponible"

    bitacora = Bitacora(
        id_usuario=current_user.id_usuario,
        accion="confirmar_llegada_manual_admin",
        descripcion=(
            f"Llegada del viaje #{viaje.id_viaje} ({viaje.origen} → {viaje.destino}, "
            f"conductor {viaje.conductor.nombre}) confirmada manualmente por "
            f"{current_user.nombre} (estado previo: '{estado_anterior}')."
        ),
        tabla_afectada="viajes",
        registro_id=viaje.id_viaje,
    )
    db.session.add(bitacora)
    db.session.commit()

    flash(f"Llegada del viaje #{viaje.id_viaje} confirmada manualmente.", "success")
    return redirect(url_for("admin.viajes_lista"))


@admin.route("/viajes/<int:id_viaje>/cerrar-forzado", methods=["POST"])
@admin_required
def viajes_cerrar_forzado(id_viaje):
    """Cierra un viaje atorado o con datos inconsistentes que no se puede
    resolver de otra forma. A diferencia de la confirmación (normal o
    manual), el estado queda en 'cerrado_admin' -no 'completado'- para dejar
    constancia de que el cierre fue anómalo, no una llegada real."""
    viaje = Viaje.query.get_or_404(id_viaje)

    if viaje.estado not in VIAJE_ESTADOS_CERRABLES_POR_ADMIN:
        flash("Ese viaje ya no está en curso.", "error")
        return redirect(url_for("admin.viajes_lista"))

    estado_anterior = viaje.estado
    viaje.hora_llegada = datetime.now()
    viaje.estado = "cerrado_admin"
    viaje.vehiculo.estado = "disponible"

    bitacora = Bitacora(
        id_usuario=current_user.id_usuario,
        accion="cerrar_viaje_forzado_admin",
        descripcion=(
            f"Viaje #{viaje.id_viaje} ({viaje.origen} → {viaje.destino}, "
            f"conductor {viaje.conductor.nombre}) cerrado de forma forzada por "
            f"{current_user.nombre} (estado previo: '{estado_anterior}')."
        ),
        tabla_afectada="viajes",
        registro_id=viaje.id_viaje,
    )
    db.session.add(bitacora)
    db.session.commit()

    flash(f"Viaje #{viaje.id_viaje} cerrado de forma forzada.", "success")
    return redirect(url_for("admin.viajes_lista"))


def _conductores_disponibles_para_programar(excluir_viaje_id=None):
    """Conductores con cuenta de acceso activa, licencia vigente y sin una
    ruta sin cerrar, aptos para que el admin les asigne una ruta nueva
    (RF-3). Un conductor con licencia vencida no debe ni aparecer en el
    selector -RF-6.2 se refuerza aquí, además de al iniciar el viaje (ver
    conductor.viajes_iniciar). El filtro de "sin ruta sin cerrar" replica el
    mismo criterio que el conflicto validado al guardar en viajes_programar,
    para que el admin no se entere del choque hasta el POST.

    El filtro de disponibilidad usa Usuario.activo (el estado de la cuenta),
    no Conductor.activo: son columnas distintas y cuentas_cambiar_estado
    (desactivar cuenta desde el admin) solo toca la primera. Es el mismo
    criterio que conductores_lista ya usa para mostrar "Activa"/"Inactiva",
    así que un conductor con la cuenta desactivada no debe seguir
    apareciendo aquí aunque Conductor.activo no se haya tocado.

    excluir_viaje_id (opcional): ignora esa ruta puntual al revisar viajes
    sin cerrar. Usado por viajes_editar_programada (auto-exclusión, RF-3):
    el conductor ya asignado a la ruta que se está editando no es un
    conflicto consigo mismo, así que debe seguir apareciendo como
    seleccionable. viajes_programar (alta) sigue llamando esta función sin
    el parámetro, así que su comportamiento no cambia.
    """
    query_sin_cerrar = Viaje.query.filter(Viaje.estado.in_(VIAJE_ESTADOS_SIN_CERRAR))
    if excluir_viaje_id is not None:
        query_sin_cerrar = query_sin_cerrar.filter(Viaje.id_viaje != excluir_viaje_id)
    ids_con_viaje_sin_cerrar = query_sin_cerrar.with_entities(Viaje.id_conductor)

    return [
        c
        for c in Conductor.query.join(Usuario, Usuario.id_conductor == Conductor.id_conductor)
        .filter(Usuario.activo.is_(True))
        .filter(~Conductor.id_conductor.in_(ids_con_viaje_sin_cerrar))
        .order_by(Conductor.nombre)
        .all()
        if c.licencia_vigente()
    ]


def _vehiculos_disponibles_para_programar(excluir_viaje_id=None):
    """Vehículos aptos para asignar a una ruta: solo 'disponible'.

    excluir_viaje_id (opcional): además del criterio anterior, incluye el
    vehículo ya asignado a esa ruta puntual aunque su columna `estado` ya
    no sea 'disponible' (ej. el mecánico lo mandó a 'en_taller' mientras
    estaba reservado). Mismo espíritu de auto-exclusión que
    _conductores_disponibles_para_programar, usado por
    viajes_editar_programada; viajes_programar (alta) sigue llamando esta
    función sin el parámetro, así que su comportamiento no cambia.
    """
    filtro = Vehiculo.estado == "disponible"
    if excluir_viaje_id is not None:
        id_vehiculo_actual = db.session.query(Viaje.id_vehiculo).filter_by(
            id_viaje=excluir_viaje_id
        ).scalar()
        if id_vehiculo_actual is not None:
            filtro = db.or_(filtro, Vehiculo.id_vehiculo == id_vehiculo_actual)
    return Vehiculo.query.filter(filtro).order_by(Vehiculo.placas).all()


def _contexto_selectores_ruta(excluir_viaje_id=None):
    """Conductores y vehículos disponibles para los selectores de
    viajes_programar/viajes_editar_programada, junto con si cada lista
    salió vacía. Calculado aquí (no en la plantilla) para que tanto el
    mensaje explicativo de selector vacío como el atributo `disabled` del
    botón de guardar lean el mismo par de banderas, en vez de repetir
    `not conductores`/`not vehiculos` en varios lugares del template."""
    conductores = _conductores_disponibles_para_programar(excluir_viaje_id=excluir_viaje_id)
    vehiculos = _vehiculos_disponibles_para_programar(excluir_viaje_id=excluir_viaje_id)
    return {
        "conductores": conductores,
        "vehiculos": vehiculos,
        "sin_conductores": not conductores,
        "sin_vehiculos": not vehiculos,
    }


def _eta_min_programar():
    """Ayuda de UX para el atributo min del input datetime-local de ETA (ver
    validar_datos_viaje, que es la validación real): mismo umbral de 1
    minuto, calculado en el momento de renderizar la página."""
    return (datetime.now() + timedelta(minutes=1)).strftime("%Y-%m-%dT%H:%M")


@admin.route("/viajes/programar", methods=["GET", "POST"])
@admin_required
def viajes_programar():
    """Programa una ruta nueva: el admin elige conductor + vehículo +
    origen/destino/ETA, y el viaje queda en estado 'programado' hasta que el
    conductor lo inicie con un clic desde su dashboard (RF-3).
    """
    if request.method == "POST":
        contexto_formulario = {
            **_contexto_selectores_ruta(),
            "formulario": request.form,
            "eta_min": _eta_min_programar(),
        }

        conductor = Conductor.query.get(request.form.get("id_conductor", type=int))
        if conductor is None or not conductor.cuenta_activa() or not conductor.licencia_vigente():
            flash("Selecciona un conductor activo con licencia vigente.", "error")
            return render_template("admin/viajes/programar.html", **contexto_formulario)

        vehiculo = Vehiculo.query.get(request.form.get("id_vehiculo", type=int))
        if vehiculo is None or vehiculo.estado != "disponible":
            flash("Selecciona un vehículo disponible.", "error")
            return render_template("admin/viajes/programar.html", **contexto_formulario)

        origen, destino, eta, error = validar_datos_viaje(
            request.form.get("origen"), request.form.get("destino"), request.form.get("eta")
        )
        if error:
            flash(error, "error")
            return render_template("admin/viajes/programar.html", **contexto_formulario)

        # No basta con que la licencia esté vigente hoy (ya validado arriba):
        # una ruta programada con días de anticipación puede tener una ETA
        # posterior a la fecha de vencimiento de la licencia, y el conductor
        # llegaría a su destino ya sin licencia vigente (RF-6.2 extendido a
        # la ETA, no solo a "hoy").
        if not conductor.licencia_vigente_en(eta.date()):
            flash(
                "La licencia del conductor seleccionado vence antes de la ETA de esta ruta.",
                "error",
            )
            return render_template("admin/viajes/programar.html", **contexto_formulario)

        # Mismo espíritu que RF-3.6: ni el conductor ni el vehículo pueden
        # tener ya otro viaje sin cerrar (programado, activo, en alerta o en
        # emergencia) antes de programarles uno nuevo.
        conflicto = Viaje.query.filter(
            db.or_(
                Viaje.id_conductor == conductor.id_conductor,
                Viaje.id_vehiculo == vehiculo.id_vehiculo,
            ),
            Viaje.estado.in_(VIAJE_ESTADOS_SIN_CERRAR),
        ).first()
        if conflicto is not None:
            flash("El conductor o el vehículo seleccionados ya tienen un viaje sin cerrar.", "error")
            return render_template("admin/viajes/programar.html", **contexto_formulario)

        # hora_salida es NOT NULL en el modelo y esa restricción no se toca:
        # se usa la hora de programación como valor temporal, que se
        # sobreescribe con la hora real cuando el conductor inicia el viaje
        # (ver conductor.viajes_iniciar).
        viaje = Viaje(
            id_conductor=conductor.id_conductor,
            id_vehiculo=vehiculo.id_vehiculo,
            origen=origen,
            destino=destino,
            hora_salida=datetime.now(),
            eta=eta,
            estado="programado",
        )

        try:
            db.session.add(viaje)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("El conductor o el vehículo seleccionados ya tienen un viaje sin cerrar.", "error")
            return render_template("admin/viajes/programar.html", **contexto_formulario)

        usuario_conductor = Usuario.query.filter_by(id_conductor=conductor.id_conductor).first()
        if usuario_conductor is not None and usuario_conductor.email:
            enviar_correo(
                usuario_conductor.email,
                "Nueva ruta asignada — S.A.D.E.",
                (
                    f"Hola {conductor.nombre},\n\n"
                    "Se te asignó una nueva ruta:\n"
                    f"Origen: {origen}\n"
                    f"Destino: {destino}\n"
                    f"ETA: {eta.strftime('%d/%m/%Y %H:%M')}\n\n"
                    "Inicia sesión en S.A.D.E. y da clic en \"Iniciar viaje\" cuando estés listo para salir."
                ),
            )
            # Un fallo de envío (correo mal configurado, SMTP caído) no debe
            # romper la programación de la ruta: ya quedó guardada en BD, que
            # es la fuente de verdad; el correo es solo un aviso adicional.

        bitacora = Bitacora(
            id_usuario=current_user.id_usuario,
            accion="ruta_programada",
            descripcion=(
                f"Ruta programada para el conductor {conductor.nombre} con el vehículo "
                f"{vehiculo.placas} ({origen} → {destino}, ETA {eta.strftime('%d/%m/%Y %H:%M')}) "
                f"por {current_user.nombre}."
            ),
            tabla_afectada="viajes",
            registro_id=viaje.id_viaje,
        )
        db.session.add(bitacora)
        db.session.commit()

        flash(f"Ruta programada correctamente para {conductor.nombre}.", "success")
        return redirect(url_for("admin.viajes_programadas"))

    return render_template(
        "admin/viajes/programar.html",
        **_contexto_selectores_ruta(),
        formulario=None,
        eta_min=_eta_min_programar(),
    )


@admin.route("/viajes/programadas")
@admin_required
def viajes_programadas():
    """Lista las rutas programadas por el admin que siguen pendientes de que
    el conductor las inicie (RF-3). Pasa 'ahora' para que el template pinte
    el badge de ETA vencida sin depender de que el scheduler ya haya
    corrido (ver scheduler.revisar_rutas_no_iniciadas)."""
    rutas = (
        Viaje.query.options(joinedload(Viaje.conductor), joinedload(Viaje.vehiculo))
        .filter_by(estado="programado")
        .order_by(Viaje.eta)
        .all()
    )
    return render_template("admin/viajes/programadas.html", rutas=rutas, ahora=datetime.now())


@admin.route("/viajes/<int:id_viaje>/cancelar-programada", methods=["POST"])
@admin_required
def viajes_cancelar_programada(id_viaje):
    """Cancela una ruta programada que el conductor todavía no inició.

    Si la ETA no había vencido (no existe Alerta 'ruta_no_iniciada' para
    ella), nada más depende del viaje y se borra el registro directamente,
    igual que antes.

    Si SÍ existe esa alerta, el viaje ya no se puede borrar:
    alertas.id_viaje -> viajes.id_viaje tiene FK con ON DELETE NO ACTION en
    la BD real, así que db.session.delete(viaje) lanzaría IntegrityError. En
    ese caso se conserva el Viaje (estado 'cerrado_admin', el mismo que usa
    viajes_cerrar_forzado para "cerrado por el admin fuera del flujo
    normal"; a diferencia de ese caso, aquí el vehículo nunca dejó de estar
    'disponible' porque la ruta nunca arrancó, así que no se toca) y se
    marca la alerta como atendida, para no perder en el Historial la
    trazabilidad de que se resolvió por cancelación manual (ver
    _cierre_para_ruta_no_iniciada).
    """
    viaje = Viaje.query.get_or_404(id_viaje)

    if viaje.estado != "programado":
        flash("Esa ruta ya no está pendiente de iniciar.", "error")
        return redirect(url_for("admin.viajes_programadas"))

    descripcion = (
        f"Ruta programada #{viaje.id_viaje} ({viaje.origen} → {viaje.destino}, "
        f"conductor {viaje.conductor.nombre}) cancelada por {current_user.nombre} "
        "antes de que el conductor la iniciara."
    )

    alerta_ruta_no_iniciada = Alerta.query.filter_by(
        id_viaje=viaje.id_viaje, tipo="ruta_no_iniciada"
    ).first()

    if alerta_ruta_no_iniciada is not None:
        ahora = datetime.now()
        alerta_ruta_no_iniciada.atendida = True
        alerta_ruta_no_iniciada.atendida_en = ahora
        viaje.estado = "cerrado_admin"
        viaje.hora_llegada = ahora
    else:
        db.session.delete(viaje)

    db.session.commit()

    bitacora = Bitacora(
        id_usuario=current_user.id_usuario,
        accion="ruta_programada_cancelada",
        descripcion=descripcion,
        tabla_afectada="viajes",
        registro_id=id_viaje,
    )
    db.session.add(bitacora)
    db.session.commit()

    flash("Ruta programada cancelada.", "success")
    return redirect(url_for("admin.viajes_programadas"))


@admin.route("/viajes/<int:id_viaje>/editar-programada", methods=["GET", "POST"])
@admin_required
def viajes_editar_programada(id_viaje):
    """Edita conductor, vehículo, origen/destino/ETA de una ruta programada
    que el conductor todavía no inició (RF-3): antes la única forma de
    corregir un dato mal capturado era cancelar y reprogramar desde cero,
    riesgoso dado el margen mínimo de solo 1 minuto que ya tiene la ETA (ver
    validar_datos_viaje).

    Solo aplica con estado 'programado' -en GET y en POST, por si el
    conductor la inicia entre ambas peticiones-, igual que
    viajes_cancelar_programada.

    AUTO-EXCLUSIÓN (mismo espíritu que el aviso de placa duplicada al
    editar el vehículo que ya tiene esa misma placa): el conductor y el
    vehículo YA asignados a esta ruta no son un conflicto consigo mismos,
    así que _conductores_disponibles_para_programar /
    _vehiculos_disponibles_para_programar reciben excluir_viaje_id=id_viaje
    para seguir mostrándolos como seleccionables, y el chequeo de conflicto
    de más abajo excluye este mismo viaje al buscar otro viaje sin cerrar.
    """
    viaje = Viaje.query.get_or_404(id_viaje)

    if viaje.estado != "programado":
        flash("Esa ruta ya no está pendiente de iniciar; no se puede editar.", "error")
        return redirect(url_for("admin.viajes_programadas"))

    if request.method == "POST":
        contexto_formulario = {
            **_contexto_selectores_ruta(excluir_viaje_id=id_viaje),
            "formulario": request.form,
            "eta_min": _eta_min_programar(),
            "viaje": viaje,
        }

        conductor = Conductor.query.get(request.form.get("id_conductor", type=int))
        if conductor is None or not conductor.cuenta_activa() or not conductor.licencia_vigente():
            flash("Selecciona un conductor activo con licencia vigente.", "error")
            return render_template("admin/viajes/editar_programada.html", **contexto_formulario)

        vehiculo = Vehiculo.query.get(request.form.get("id_vehiculo", type=int))
        # Auto-exclusión: el vehículo ya asignado a esta ruta es válido
        # aunque su columna `estado` ya no sea 'disponible' (ver
        # _vehiculos_disponibles_para_programar).
        if vehiculo is None or (
            vehiculo.estado != "disponible" and vehiculo.id_vehiculo != viaje.id_vehiculo
        ):
            flash("Selecciona un vehículo disponible.", "error")
            return render_template("admin/viajes/editar_programada.html", **contexto_formulario)

        origen, destino, eta, error = validar_datos_viaje(
            request.form.get("origen"), request.form.get("destino"), request.form.get("eta")
        )
        if error:
            flash(error, "error")
            return render_template("admin/viajes/editar_programada.html", **contexto_formulario)

        # Mismo motivo que en viajes_programar: la ETA editada puede caer
        # después del vencimiento de la licencia aunque siga vigente hoy.
        if not conductor.licencia_vigente_en(eta.date()):
            flash(
                "La licencia del conductor seleccionado vence antes de la ETA de esta ruta.",
                "error",
            )
            return render_template("admin/viajes/editar_programada.html", **contexto_formulario)

        # Mismo criterio que viajes_programar, excluyendo esta misma ruta:
        # ni el conductor ni el vehículo nuevos pueden estar ya
        # comprometidos con OTRA ruta sin cerrar (programada, activa, en
        # alerta o en emergencia).
        conflicto = Viaje.query.filter(
            Viaje.id_viaje != viaje.id_viaje,
            db.or_(
                Viaje.id_conductor == conductor.id_conductor,
                Viaje.id_vehiculo == vehiculo.id_vehiculo,
            ),
            Viaje.estado.in_(VIAJE_ESTADOS_SIN_CERRAR),
        ).first()
        if conflicto is not None:
            flash("El conductor o el vehículo seleccionados ya tienen un viaje sin cerrar.", "error")
            return render_template("admin/viajes/editar_programada.html", **contexto_formulario)

        cambios = []
        if conductor.id_conductor != viaje.id_conductor:
            cambios.append(f"conductor de '{viaje.conductor.nombre}' a '{conductor.nombre}'")
        if vehiculo.id_vehiculo != viaje.id_vehiculo:
            cambios.append(f"vehículo de '{viaje.vehiculo.placas}' a '{vehiculo.placas}'")
        if origen != viaje.origen or destino != viaje.destino:
            cambios.append(f"ruta de '{viaje.origen} → {viaje.destino}' a '{origen} → {destino}'")
        if eta != viaje.eta:
            cambios.append(
                f"ETA de '{viaje.eta.strftime('%d/%m/%Y %H:%M')}' a '{eta.strftime('%d/%m/%Y %H:%M')}'"
            )

        viaje.id_conductor = conductor.id_conductor
        viaje.id_vehiculo = vehiculo.id_vehiculo
        viaje.origen = origen
        viaje.destino = destino
        viaje.eta = eta

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("El conductor o el vehículo seleccionados ya tienen un viaje sin cerrar.", "error")
            return render_template("admin/viajes/editar_programada.html", **contexto_formulario)

        if cambios:
            bitacora = Bitacora(
                id_usuario=current_user.id_usuario,
                accion="ruta_programada_editada",
                descripcion=(
                    f"Ruta programada #{viaje.id_viaje} editada por {current_user.nombre}: "
                    + "; ".join(cambios) + "."
                ),
                tabla_afectada="viajes",
                registro_id=viaje.id_viaje,
            )
            db.session.add(bitacora)
            db.session.commit()

        flash("Ruta programada actualizada correctamente.", "success")
        return redirect(url_for("admin.viajes_programadas"))

    return render_template(
        "admin/viajes/editar_programada.html",
        **_contexto_selectores_ruta(excluir_viaje_id=id_viaje),
        formulario={
            "id_conductor": viaje.id_conductor,
            "id_vehiculo": viaje.id_vehiculo,
            "origen": viaje.origen,
            "destino": viaje.destino,
            "eta": viaje.eta.strftime("%Y-%m-%dT%H:%M"),
        },
        eta_min=_eta_min_programar(),
        viaje=viaje,
    )


def _detalle_alerta(alerta):
    """Arma los datos del panel de detalle de una alerta (RF-3.5/RF-4): la
    Emergencia asociada, que solo aplica a tipo='panico' -las de tipo
    'retraso' o 'licencia_vencida' nunca disparan Twilio (ver
    app/controllers/emergencia.py)-, para mostrar el estado real de envío
    de WhatsApp/SMS en vez de inventar uno.
    """
    emergencia = None
    if alerta.tipo == "panico" and alerta.viaje is not None:
        emergencia = (
            Emergencia.query.filter_by(id_viaje=alerta.viaje.id_viaje)
            .order_by(Emergencia.activada_en.desc())
            .first()
        )
    return {"emergencia": emergencia}


_TIPOS_ALERTA_VALIDOS = (
    "retraso",
    "panico",
    "licencia_vencida",
    "asistencia_mecanica",
    "incidencia_trafico",
    "ruta_no_iniciada",
)


@admin.route("/alertas")
@admin_required
def alertas_lista():
    """Vista de dos columnas: alertas sin atender a la izquierda (RF-3.5,
    igual que antes) y el detalle de la seleccionada a la derecha
    (?id_alerta=<id>; si falta o no corresponde a una alerta sin atender, se
    usa la primera de la lista), más 4 tarjetas de resumen.

    ?tipo=<tipo> (repetible, ej. ?tipo=asistencia_mecanica&tipo=incidencia_trafico)
    reduce la lista a uno o más tipos -usado por la tarjeta "Emergencias" del
    dashboard, que agrupa mecánica + tráfico-; sin el parámetro (o con
    valores que no correspondan a ningún tipo válido) se conserva el
    comportamiento default: todas las alertas sin atender.
    """
    tipos_filtro = [t for t in request.args.getlist("tipo") if t in _TIPOS_ALERTA_VALIDOS]

    query_alertas = (
        Alerta.query.options(joinedload(Alerta.conductor), joinedload(Alerta.viaje))
        .filter_by(atendida=False)
    )
    if tipos_filtro:
        query_alertas = query_alertas.filter(Alerta.tipo.in_(tipos_filtro))
    alertas = query_alertas.order_by(Alerta.generada_en.desc()).all()

    id_alerta_param = request.args.get("id_alerta", type=int)
    alerta_seleccionada = None
    if id_alerta_param is not None:
        alerta_seleccionada = next(
            (a for a in alertas if a.id_alerta == id_alerta_param), None
        )
    if alerta_seleccionada is None and alertas:
        alerta_seleccionada = alertas[0]

    detalle = _detalle_alerta(alerta_seleccionada) if alerta_seleccionada else None

    hoy = date.today()
    inicio_hoy = datetime(hoy.year, hoy.month, hoy.day)
    inicio_mes = datetime(hoy.year, hoy.month, 1)

    conteos = {
        "hoy": Alerta.query.filter(Alerta.generada_en >= inicio_hoy).count(),
        "sin_atender": Alerta.query.filter_by(atendida=False).count(),
        "retrasos_mes": Alerta.query.filter(
            Alerta.tipo == "retraso", Alerta.generada_en >= inicio_mes
        ).count(),
        "emergencias_mes": Alerta.query.filter(
            Alerta.tipo == "panico", Alerta.generada_en >= inicio_mes
        ).count(),
    }

    return render_template(
        "admin/alertas/lista.html",
        alertas=alertas,
        alerta_seleccionada=alerta_seleccionada,
        detalle=detalle,
        conteos=conteos,
        tipos_filtro=tipos_filtro,
    )


@admin.route("/reportes-averia")
@admin_required
def reportes_averia_lista():
    """Lista los reportes de avería sin archivar registrados por mecánica
    (RF-5.4). Por default solo muestra archivado_en IS NULL: un reporte
    archivado no se borra (ver reportes_averia_archivar), solo se oculta de
    esta lista -- el Historial (_construir_eventos_historial) lo sigue
    encontrando sin importar este filtro, ver esa función.

    No hay infraestructura de tiempo real: esta vista simplemente consulta la
    BD en cada carga, igual que Vehículos y Viajes activos.
    """
    reportes = (
        ReporteAveria.query.options(
            joinedload(ReporteAveria.vehiculo),
            joinedload(ReporteAveria.viaje),
            joinedload(ReporteAveria.usuario),
        )
        .filter(ReporteAveria.archivado_en.is_(None))
        .order_by(ReporteAveria.registrado_en.desc())
        .all()
    )
    return render_template("admin/reportes_averia/lista.html", reportes=reportes)


@admin.route("/reportes-averia/archivar", methods=["POST"])
@admin_required
def reportes_averia_archivar():
    """Archiva (no borra) los reportes de avería más viejos que N meses
    (RF-5.4, punto 1 de la profesora): oculta reportes viejos de la lista
    operativa sin perder el dato para el Historial ni para auditoría.

    El corte de "N meses" se aproxima como N*30 días -- no hay una librería
    de fechas calendáricas en el proyecto (ver requirements.txt) y para este
    filtro, pensado para limpieza visual y no para un reporte contable, la
    aproximación es suficiente.
    """
    meses_validos = {4, 8, 12}
    meses = request.form.get("meses", type=int)
    if meses not in meses_validos:
        flash("Selecciona un periodo válido para archivar (4, 8 o 12 meses).", "error")
        return redirect(url_for("admin.reportes_averia_lista"))

    fecha_corte = datetime.now() - timedelta(days=meses * 30)
    reportes_a_archivar = ReporteAveria.query.filter(
        ReporteAveria.registrado_en < fecha_corte,
        ReporteAveria.archivado_en.is_(None),
    ).all()

    ahora = datetime.now()
    for reporte in reportes_a_archivar:
        reporte.archivado_en = ahora

    bitacora = Bitacora(
        id_usuario=current_user.id_usuario,
        accion="archivar_reportes_averia",
        descripcion=(
            f"{len(reportes_a_archivar)} reporte(s) de avería anteriores a "
            f"{meses} meses ({fecha_corte.strftime('%d/%m/%Y')}) archivados por "
            f"{current_user.nombre}."
        ),
        tabla_afectada="reportes_averia",
        # registro_id es NOT NULL pero esta acción afecta a varios reportes a
        # la vez, no a uno solo: se usa 0 como valor centinela (ningún
        # id_reporte real es 0) porque no hay un único registro al que
        # apuntar; la descripción ya trae el conteo exacto.
        registro_id=0,
    )
    db.session.add(bitacora)
    db.session.commit()

    flash(f"Se archivaron {len(reportes_a_archivar)} reporte(s) de avería.", "success")
    return redirect(url_for("admin.reportes_averia_lista"))


def _tipo_y_cierre_de_viaje_cerrado(viaje, reportes_por_viaje, forzado_por_viaje, manual_por_viaje):
    """Deriva (slug_tipo, cierre) para un Viaje ya cerrado ('completado' o
    'cerrado_admin'), revisando en orden de prioridad:

    1. Si el mecánico registró un ReporteAveria para ese viaje (con el
       checkbox de "cambiar a en_taller" marcado, cerró el viaje) -> avería.
       Se revisa PRIMERO porque ese cierre también deja estado='cerrado_admin',
       igual que el cierre forzado del admin, y hay que distinguirlos.
    2. Si hay Bitácora de cerrar_viaje_forzado_admin para ese viaje -> forzado.
    3. Si hay Bitácora de confirmar_llegada_manual_admin para ese viaje -> manual.
    4. Si no aplica ninguna de las anteriores -> cierre normal del conductor
       (viajes_confirmar_llegada en conductor.py no deja registro en Bitácora).
    """
    if viaje.id_viaje in reportes_por_viaje:
        return "averia", "Mecánico"
    if viaje.id_viaje in forzado_por_viaje:
        return "cerrado_forzado", "Manual (admin)"
    if viaje.id_viaje in manual_por_viaje:
        return "completado", "Manual (admin)"
    return "completado", "Conductor"


def _cierre_para_panico(alerta, reportes_por_viaje, forzado_por_viaje, manual_por_viaje):
    """Cierre de una Alerta tipo='panico', derivado del cierre de su Viaje
    asociado con la misma prioridad de _tipo_y_cierre_de_viaje_cerrado().

    Si el viaje de la emergencia todavía no se cerró (sigue
    activo/alerta/emergencia), no hay nada que derivar todavía: se reporta
    'Pendiente' en vez de adivinar un cierre que no ha ocurrido.
    """
    viaje = alerta.viaje
    if viaje is None or viaje.estado not in ("completado", "cerrado_admin"):
        return "Pendiente"
    _, cierre = _tipo_y_cierre_de_viaje_cerrado(viaje, reportes_por_viaje, forzado_por_viaje, manual_por_viaje)
    return cierre


def _cierre_para_ruta_no_iniciada(alerta, cancelada_por_viaje):
    """Cierre de una Alerta tipo='ruta_no_iniciada', sin pantalla de
    "atender" dedicada (ver admin.viajes_cancelar_programada):

    - El admin canceló la ruta -> 'Manual (admin)'. Se revisa primero: ese
      viaje queda en estado 'cerrado_admin' pero se excluye del bucket de
      viajes_cerrados más abajo, precisamente para que esta sea la ÚNICA
      fila de Historial para ese evento, no una duplicada y mal etiquetada.
    - 'Conductor' -> alerta.atendida en True sin cancelación del admin. Ya
      no puede volver a ocurrir (conductor.viajes_iniciar ahora rechaza de
      forma permanente iniciar una ruta con ETA vencida, así que el inicio
      tardío dejó de resolver esta alerta); se conserva la rama solo para
      no romper el Historial de alertas creadas antes de ese cambio.
    - Si no ha pasado ninguna de las dos todavía -> 'Pendiente'.
    """
    if alerta.id_viaje in cancelada_por_viaje:
        return "Manual (admin)"
    if alerta.atendida:
        return "Conductor"
    return "Pendiente"


def _construir_eventos_historial():
    """Arma la lista normalizada del Historial (RF de auditoría): une
    Alertas (retraso/pánico/licencia_vencida) y Viajes cerrados
    (completado/cerrado_admin) en un solo tipo de evento, ordenada por
    fecha descendente. No hay hoy ninguna Alerta tipo='licencia_vencida'
    real en el sistema (el check-in solo bloquea con un flash, RF-6.2, sin
    crear fila); el código la contempla por si alguna vez existe, pero su
    ausencia no debe romper ni vaciar el resto del historial.
    """
    reportes_por_viaje = {r.id_viaje: r for r in ReporteAveria.query.all()}
    forzado_por_viaje = {
        b.registro_id: b
        for b in Bitacora.query.filter_by(accion="cerrar_viaje_forzado_admin", tabla_afectada="viajes").all()
    }
    manual_por_viaje = {
        b.registro_id: b
        for b in Bitacora.query.filter_by(accion="confirmar_llegada_manual_admin", tabla_afectada="viajes").all()
    }
    cancelada_por_viaje = {
        b.registro_id: b
        for b in Bitacora.query.filter_by(accion="ruta_programada_cancelada", tabla_afectada="viajes").all()
    }

    eventos = []

    alertas = (
        Alerta.query.options(joinedload(Alerta.conductor), joinedload(Alerta.viaje))
        .filter(
            Alerta.tipo.in_(
                [
                    "retraso",
                    "panico",
                    "licencia_vencida",
                    "asistencia_mecanica",
                    "incidencia_trafico",
                    "ruta_no_iniciada",
                ]
            )
        )
        .all()
    )
    for alerta in alertas:
        if alerta.tipo == "retraso":
            slug, cierre = "retraso", "Automático"
            ruta = f"{alerta.viaje.origen} → {alerta.viaje.destino}" if alerta.viaje else "—"
            id_vehiculo = alerta.viaje.id_vehiculo if alerta.viaje else None
        elif alerta.tipo == "panico":
            slug = "panico"
            cierre = _cierre_para_panico(alerta, reportes_por_viaje, forzado_por_viaje, manual_por_viaje)
            ruta = f"{alerta.viaje.origen} → {alerta.viaje.destino}" if alerta.viaje else "—"
            id_vehiculo = alerta.viaje.id_vehiculo if alerta.viaje else None
        elif alerta.tipo in ("asistencia_mecanica", "incidencia_trafico"):
            # El conductor es siempre quien las genera (RF-5 para la
            # mecánica; informativa para el admin en el caso de tráfico): a
            # diferencia de panico/viajes cerrados, no hay nada que derivar
            # del cierre del viaje asociado.
            slug, cierre = alerta.tipo, "Conductor"
            ruta = f"{alerta.viaje.origen} → {alerta.viaje.destino}" if alerta.viaje else "—"
            id_vehiculo = alerta.viaje.id_vehiculo if alerta.viaje else None
        elif alerta.tipo == "ruta_no_iniciada":
            slug = "ruta_no_iniciada"
            cierre = _cierre_para_ruta_no_iniciada(alerta, cancelada_por_viaje)
            ruta = f"{alerta.viaje.origen} → {alerta.viaje.destino}" if alerta.viaje else "—"
            id_vehiculo = alerta.viaje.id_vehiculo if alerta.viaje else None
        else:  # licencia_vencida (contemplado por esquema; hoy nunca se genera)
            slug, cierre = "licencia_vencida", "Sistema"
            ruta = "—"
            id_vehiculo = None

        eventos.append(
            {
                "fecha": alerta.generada_en,
                "tipo_slug": slug,
                "conductor": alerta.conductor.nombre,
                "id_conductor": alerta.id_conductor,
                "ruta": ruta,
                "id_vehiculo": id_vehiculo,
                "cierre": cierre,
            }
        )

    viajes_cerrados = (
        Viaje.query.options(joinedload(Viaje.conductor), joinedload(Viaje.vehiculo))
        .filter(Viaje.estado.in_(["completado", "cerrado_admin"]))
        .filter(Viaje.id_viaje.notin_(cancelada_por_viaje.keys()))
        .all()
    )
    for viaje in viajes_cerrados:
        slug, cierre = _tipo_y_cierre_de_viaje_cerrado(viaje, reportes_por_viaje, forzado_por_viaje, manual_por_viaje)
        eventos.append(
            {
                "fecha": viaje.hora_llegada,
                "tipo_slug": slug,
                "conductor": viaje.conductor.nombre,
                "id_conductor": viaje.id_conductor,
                "ruta": f"{viaje.origen} → {viaje.destino}",
                "id_vehiculo": viaje.id_vehiculo,
                "cierre": cierre,
            }
        )

    # Ordena por fecha descendente. En teoria hora_llegada siempre esta
    # presente para un viaje 'completado'/'cerrado_admin' (todas las rutas
    # que producen esos estados la fijan), pero no hay una restriccion de BD
    # que lo garantice y ya existen filas reales mas viejas sin ese dato:
    # se tratan como la fecha mas antigua posible en vez de romper el orden.
    eventos.sort(key=lambda e: e["fecha"] or datetime.min, reverse=True)
    for evento in eventos:
        etiqueta, badge, endpoint = TIPO_INFO_HISTORIAL[evento["tipo_slug"]]
        evento["tipo_etiqueta"] = etiqueta
        evento["badge"] = badge
        evento["url_detalle"] = url_for(endpoint)
    return eventos


def _exportar_historial_csv(eventos):
    """Genera el CSV de descarga con las mismas columnas visibles en la tabla."""
    buffer = io.StringIO()
    escritor = csv.writer(buffer)
    escritor.writerow(["Tipo de evento", "Conductor", "Ruta", "Fecha y hora", "Cierre"])
    for evento in eventos:
        escritor.writerow(
            [
                evento["tipo_etiqueta"],
                evento["conductor"],
                evento["ruta"],
                evento["fecha"].strftime("%d/%m/%Y %H:%M") if evento["fecha"] else "",
                evento["cierre"],
            ]
        )

    # utf-8-sig antepone el BOM que Excel necesita para reconocer la
    # codificación automáticamente; sin él, Excel asume Latin-1 y los
    # acentos/flechas del contenido se muestran corruptos.
    respuesta = Response(buffer.getvalue().encode("utf-8-sig"), mimetype="text/csv")
    respuesta.headers["Content-Disposition"] = "attachment; filename=historial_sade.csv"
    return respuesta


COLUMNAS_HISTORIAL = ("Tipo de evento", "Conductor", "Ruta", "Fecha y hora", "Cierre")
ANCHO_COLUMNA_MAX_XLSX = 50


def _exportar_historial_xlsx(eventos):
    """Genera el Excel de descarga con las mismas columnas visibles en la
    tabla y en _exportar_historial_csv(), con formato: encabezado con fondo
    corporativo, bordes en las celdas con datos y columnas congeladas."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Historial"

    relleno_encabezado = PatternFill(start_color="17423F", end_color="17423F", fill_type="solid")
    fuente_encabezado = Font(color="FFFFFF", bold=True)
    borde_delgado = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    hoja.append(COLUMNAS_HISTORIAL)
    for celda in hoja[1]:
        celda.fill = relleno_encabezado
        celda.font = fuente_encabezado

    for evento in eventos:
        hoja.append(
            [
                evento["tipo_etiqueta"],
                evento["conductor"],
                evento["ruta"],
                evento["fecha"].strftime("%d/%m/%Y %H:%M") if evento["fecha"] else "",
                evento["cierre"],
            ]
        )

    for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row):
        for celda in fila:
            celda.border = borde_delgado

    for indice, columna in enumerate(COLUMNAS_HISTORIAL, start=1):
        letra = get_column_letter(indice)
        ancho_contenido = max(
            (len(str(celda.value)) for celda in hoja[letra] if celda.value is not None),
            default=len(columna),
        )
        hoja.column_dimensions[letra].width = min(ancho_contenido + 2, ANCHO_COLUMNA_MAX_XLSX)

    hoja.freeze_panes = "A2"

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)

    respuesta = Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    respuesta.headers["Content-Disposition"] = "attachment; filename=historial_sade.xlsx"
    return respuesta


@admin.route("/historial")
@admin_required
def historial():
    """Historial de eventos: une Alertas y Viajes cerrados en una sola
    lista normalizada, con filtros opcionales, exportación a CSV y
    paginación (ver _construir_eventos_historial() para el detalle de cómo
    se normaliza y deriva cada tipo/cierre).
    """
    eventos = _construir_eventos_historial()

    tipo_filtro = request.args.get("tipo", "").strip()
    conductor_filtro = request.args.get("conductor", "").strip()
    vehiculo_filtro = request.args.get("vehiculo", "").strip()
    desde_texto = request.args.get("desde", "").strip()
    hasta_texto = request.args.get("hasta", "").strip()

    if tipo_filtro:
        eventos = [e for e in eventos if e["tipo_slug"] == tipo_filtro]
    if conductor_filtro:
        eventos = [e for e in eventos if str(e["id_conductor"]) == conductor_filtro]
    if vehiculo_filtro:
        eventos = [e for e in eventos if str(e["id_vehiculo"]) == vehiculo_filtro]

    if desde_texto:
        try:
            desde_fecha = datetime.strptime(desde_texto, "%Y-%m-%d")
            # Un evento sin fecha conocida (ver nota en _construir_eventos_historial)
            # no puede afirmarse que caiga dentro del rango: se excluye.
            eventos = [e for e in eventos if e["fecha"] is not None and e["fecha"] >= desde_fecha]
        except ValueError:
            flash("La fecha 'desde' no es válida.", "error")
            desde_texto = ""

    if hasta_texto:
        try:
            # +1 día para que 'hasta' incluya todo ese día completo.
            hasta_fecha = datetime.strptime(hasta_texto, "%Y-%m-%d") + timedelta(days=1)
            eventos = [e for e in eventos if e["fecha"] is not None and e["fecha"] < hasta_fecha]
        except ValueError:
            flash("La fecha 'hasta' no es válida.", "error")
            hasta_texto = ""

    formato = request.args.get("formato")
    if formato == "csv":
        return _exportar_historial_csv(eventos)
    if formato == "xlsx":
        return _exportar_historial_xlsx(eventos)

    filtros_actuales = {
        clave: valor
        for clave, valor in {
            "tipo": tipo_filtro,
            "conductor": conductor_filtro,
            "vehiculo": vehiculo_filtro,
            "desde": desde_texto,
            "hasta": hasta_texto,
        }.items()
        if valor
    }

    total_eventos = len(eventos)
    try:
        pagina = max(1, int(request.args.get("pagina", 1)))
    except ValueError:
        pagina = 1
    total_paginas = max(1, math.ceil(total_eventos / TAM_PAGINA_HISTORIAL))
    pagina = min(pagina, total_paginas)
    inicio = (pagina - 1) * TAM_PAGINA_HISTORIAL
    eventos_pagina = eventos[inicio : inicio + TAM_PAGINA_HISTORIAL]

    return render_template(
        "admin/historial/lista.html",
        eventos=eventos_pagina,
        tipos_evento=TIPO_INFO_HISTORIAL,
        conductores=Conductor.query.order_by(Conductor.nombre).all(),
        vehiculos=Vehiculo.query.order_by(Vehiculo.placas).all(),
        filtros_actuales=filtros_actuales,
        pagina=pagina,
        total_paginas=total_paginas,
        total_eventos=total_eventos,
    )


@admin.route("/alertas/<int:id_alerta>/atender", methods=["POST"])
@admin_required
def alertas_atender(id_alerta):
    """Marca una alerta como atendida (RF-3.5)."""
    alerta = Alerta.query.get_or_404(id_alerta)
    alerta.atendida = True
    alerta.atendida_en = datetime.now()
    db.session.commit()

    flash("Alerta marcada como atendida.", "success")
    return redirect(url_for("admin.alertas_lista"))


TAM_PAGINA_BITACORA = 20


@admin.route("/bitacora")
@admin_required
def bitacora_lista():
    """Registro técnico de auditoría (RNF-02/RNF-07): todas las acciones que
    se escriben en Bitacora desde cualquier parte del sistema (altas de
    cuenta, cambios de estado, reportes de avería, envíos de notificación,
    reseteos de contraseña, etc.), más recientes primero.

    Distinta de Historial a propósito: Historial es la vista operativa de
    eventos de viaje (alertas + viajes cerrados) para el monitoreo diario de
    la flota; Bitácora es el registro técnico de auditoría de acciones de
    usuarios sobre el sistema. No se fusionan.
    """
    query = Bitacora.query.options(joinedload(Bitacora.usuario))

    accion_filtro = request.args.get("accion", "").strip()
    usuario_filtro = request.args.get("usuario", "").strip()

    if accion_filtro:
        query = query.filter(Bitacora.accion == accion_filtro)
    if usuario_filtro:
        try:
            query = query.filter(Bitacora.id_usuario == int(usuario_filtro))
        except ValueError:
            usuario_filtro = ""

    query = query.order_by(Bitacora.fecha.desc())

    total_registros = query.count()
    try:
        pagina = max(1, int(request.args.get("pagina", 1)))
    except ValueError:
        pagina = 1
    total_paginas = max(1, math.ceil(total_registros / TAM_PAGINA_BITACORA))
    pagina = min(pagina, total_paginas)
    inicio = (pagina - 1) * TAM_PAGINA_BITACORA
    registros = query.offset(inicio).limit(TAM_PAGINA_BITACORA).all()

    acciones_disponibles = [
        fila[0]
        for fila in db.session.query(Bitacora.accion).distinct().order_by(Bitacora.accion).all()
    ]

    filtros_actuales = {
        clave: valor
        for clave, valor in {"accion": accion_filtro, "usuario": usuario_filtro}.items()
        if valor
    }

    return render_template(
        "admin/bitacora/lista.html",
        registros=registros,
        acciones_disponibles=acciones_disponibles,
        usuarios=Usuario.query.order_by(Usuario.nombre).all(),
        filtros_actuales=filtros_actuales,
        pagina=pagina,
        total_paginas=total_paginas,
        total_registros=total_registros,
    )
